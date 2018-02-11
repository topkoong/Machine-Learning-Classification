# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os

os.chdir('browsing/User8_b_1499982577272')
#openpyxl.load_workbook(filename) returns a Workbook object.
gyroWorkbook = openpyxl.load_workbook('gyro_readings.xlsx')
accWorkbook = openpyxl.load_workbook('acc_readings.xlsx')
wb = openpyxl.Workbook()
dest_filename = 'dataToBeCalculated.xlsx'
ws = wb.create_sheet(title="dataToBeCalculated")
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
gyroSheet = gyroWorkbook.get_sheet_by_name('gyro_readings')
accSheet = accWorkbook.get_sheet_by_name('acc_readings')

# The cell() method also returns a Cell object from a sheet.


def getXAxis():
    print("Reading Gyroscope: X Axis")
    print("Reading Accelerometer: X Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 101
    writeRowNum = 1
    maxRow = 38235
    while (readRowNum <= maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1
        print(writeRowNum, ": Reading Gyroscope X Axis from RowNo.", readRowNum," :", gyroSheet.cell(row = readRowNum, column = 2).value)
        print(writeRowNum, ": Reading Accelerometer X Axis from RowNo.", readRowNum," :", accSheet.cell(row = readRowNum, column = 2).value)
        ws.cell(row = writeRowNum, column = 2).value = gyroSheet.cell(row = readRowNum, column = 2).value
        ws.cell(row = writeRowNum, column = 5).value = accSheet.cell(row = readRowNum, column = 2).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1



def getYAxis():
    print("\nReading Gyroscope: Y Axis")
    print("Reading Accelerometer: Y Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 101
    writeRowNum = 1
    maxRow = 38235
    while (readRowNum <= maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1
        print(writeRowNum, ": Reading Gyroscope Y Axis from RowNo.", readRowNum," :", gyroSheet.cell(row = readRowNum, column = 3).value)
        print(writeRowNum, ": Reading Accelerometer Y Axis from RowNo.", readRowNum," :", accSheet.cell(row = readRowNum, column = 3).value)
        ws.cell(row = writeRowNum, column = 3).value = gyroSheet.cell(row = readRowNum, column = 3).value
        ws.cell(row = writeRowNum, column = 6).value = accSheet.cell(row = readRowNum, column = 3).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1


def getZAxis():
    print("\nReading Gyroscope: Z Axis")
    print("Reading Accelerometer: Z Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 101
    writeRowNum = 1
    maxRow = 38235
    while (readRowNum <= maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1

        print(writeRowNum, ": Reading Gyroscope Z Axis from RowNo.", readRowNum," :", gyroSheet.cell(row = readRowNum, column = 4).value)
        print(writeRowNum, ": Reading Accelerometer Z Axis from RowNo.", readRowNum," :", accSheet.cell(row = readRowNum, column = 4).value)
        ws.cell(row = writeRowNum, column = 4).value = gyroSheet.cell(row = readRowNum, column = 4).value
        ws.cell(row = writeRowNum, column = 7).value = accSheet.cell(row = readRowNum, column = 4).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1

getXAxis()
getYAxis()
getZAxis()
wb.save(filename = dest_filename)
