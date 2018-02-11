# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os
import numpy

os.chdir('browsing/user3_b_1499894113024')
dest_filename = 'dataToBeCalculated.xlsx'
#openpyxl.load_workbook(filename) returns a Workbook object.
workbook = openpyxl.load_workbook(filename = dest_filename)
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
sheet = workbook.get_sheet_by_name('dataToBeCalculated')

def findXMean():
    print("Finding Mean: X Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    xAxis = []
    i = 0
    while (readRowNum <= maxRow):
        #i = readRowNum - 1
        if(counter > 5):
            #readRowNum -= 2
            counter = 1
            print("\n\nX Axis Mean: %f" % numpy.mean(xAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 3, column = 6).value = numpy.mean(xAxis, dtype = numpy.float64)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 2).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        xAxis.insert(i, sheet.cell(row = writeRowNum, column = 2).value)
        i += 1

def findYMean():
    print("Finding Mean: Y Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    yAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            print("\n\nY Axis Mean: %f" % numpy.mean(yAxis, dtype = numpy.float64))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 3, column = 7).value = numpy.mean(yAxis, dtype = numpy.float64)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 3).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        yAxis.insert(i, sheet.cell(row = writeRowNum, column = 3).value)
        i += 1

def findZMean():
    print("Finding Mean: Z Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    zAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nZ Axis Mean: %f" % numpy.mean(zAxis, dtype = numpy.float64))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 3, column = 8).value = numpy.mean(zAxis, dtype = numpy.float64)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 4).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        zAxis.insert(i, sheet.cell(row = writeRowNum, column = 4).value)
        i += 1

findXMean()
findYMean()
findZMean()

workbook.save(filename = dest_filename)
