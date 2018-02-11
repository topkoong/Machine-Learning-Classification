# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os
import numpy

os.chdir('browsing/user3_b_1499894113024')
dest_filename = 'dataToBeCalculated.xlsx'
#openpyxl.load_workbook(filename) returns a Workbook object.
workbook = openpyxl.load_workbook(filename = dest_filename)
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
sheet = workbook.get_sheet_by_name('dataToBeCalculated')

def generateOutput():
    print("Finding Mean, Min, Max: X,Y,Z Axis from Gyroscope and Accelerometer readings")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    gyroXAxis = []
    gyroYAxis = []
    gyroZAxis = []
    accXAxis = []
    accYAxis = []
    accZAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nGyroscope: X Axis Mean: %f" % numpy.mean(gyroXAxis))
            print("\nGyroscope: X Axis Min: %f" % numpy.amin(gyroXAxis))
            print("\nGyroscope: X Axis Max: %f" % numpy.amax(gyroXAxis))
            print("=====================================\n")
            print("\n\nGyroscope: Y Axis Mean: %f" % numpy.mean(gyroYAxis))
            print("\nGyroscope: Y Axis Min: %f" % numpy.amin(gyroYAxis))
            print("\nGyroscope: Y Axis Max: %f" % numpy.amax(gyroYAxis))
            print("=====================================\n")
            print("\n\nGyroscope: Z Axis Mean: %f" % numpy.mean(gyroZAxis))
            print("\nGyroscope: Z Axis Min: %f" % numpy.amin(gyroZAxis))
            print("\nGyroscope: Z Axis Max: %f" % numpy.amax(gyroZAxis))
            print("\n\nAccelerometer: X Axis Mean: %f" % numpy.mean(accXAxis))
            print("\nAccelerometer: X Axis Min: %f" % numpy.amin(accXAxis))
            print("\nAccelerometer: X Axis Max: %f" % numpy.amax(accXAxis))
            print("=====================================\n")
            print("\n\nAccelerometer: Y Axis Mean: %f" % numpy.mean(accYAxis))
            print("\nAccelerometer: Y Axis Min: %f" % numpy.amin(accYAxis))
            print("\nAccelerometer: Y Axis Max: %f" % numpy.amax(accYAxis))
            print("=====================================\n")
            print("\n\nAccelerometer: Z Axis Mean: %f" % numpy.mean(accZAxis))
            print("\nAccelerometer: Z Axis Min: %f" % numpy.amin(accZAxis))
            print("\nAccelerometer: Z Axis Max: %f" % numpy.amax(accZAxis))
            outputFile = open("User3.txt", 'a')
            outputFile.write("%f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f, %f\n" % (
            numpy.mean(gyroXAxis), numpy.amin(gyroXAxis), numpy.amax(gyroXAxis),
            numpy.mean(gyroYAxis), numpy.amin(gyroYAxis), numpy.amax(gyroYAxis),
            numpy.mean(gyroZAxis), numpy.amin(gyroZAxis), numpy.amax(gyroZAxis),
            numpy.mean(accXAxis), numpy.amin(accXAxis), numpy.amax(accXAxis),
            numpy.mean(accYAxis), numpy.amin(accYAxis), numpy.amax(accYAxis),
            numpy.mean(accZAxis), numpy.amin(accZAxis), numpy.amax(accZAxis)))
            outputFile.close()

        print("Reading Gyroscope: X Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 2).value)
        print("Reading Gyroscope: Y Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 3).value)
        print("Reading Gyroscope: Z Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 4).value)
        print("Reading Accelerometer: X Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 5).value)
        print("Reading Accelerometer: Y Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 6).value)
        print("Reading Accelerometer: Z Axis from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 7).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        gyroXAxis.insert(i, sheet.cell(row = writeRowNum, column = 2).value)
        gyroYAxis.insert(i, sheet.cell(row = writeRowNum, column = 3).value)
        gyroZAxis.insert(i, sheet.cell(row = writeRowNum, column = 4).value)
        accXAxis.insert(i, sheet.cell(row = writeRowNum, column = 5).value)
        accYAxis.insert(i, sheet.cell(row = writeRowNum, column = 6).value)
        accZAxis.insert(i, sheet.cell(row = writeRowNum, column = 7).value)
        i += 1

generateOutput()

workbook.save(filename = dest_filename)
