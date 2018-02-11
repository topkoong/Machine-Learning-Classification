# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os
import numpy

os.chdir('browsing/User5_b_1499965432443')
dest_filename = 'dataToBeCalculated.xlsx'
#openpyxl.load_workbook(filename) returns a Workbook object.
workbook = openpyxl.load_workbook(filename = dest_filename)
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
sheet = workbook.get_sheet_by_name('dataToBeCalculated')

def findXMin():
    print("Finding Min: X Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    xAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nX Axis Min: %f" % numpy.amin(xAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 1, column = 9).value = numpy.amin(xAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 2).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        xAxis.insert(i, sheet.cell(row = writeRowNum, column = 2).value)
        i += 1

def findYMin():
    print("Finding Min: Y Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    yAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nY Axis Min: %f" % numpy.amin(yAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 1, column = 10).value = numpy.amin(yAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 3).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        yAxis.insert(i, sheet.cell(row = writeRowNum, column = 3).value)
        i += 1

def findZMin():
    print("Finding Min: Z Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    zAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nZ Axis Min: %f" % numpy.amin(zAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 1, column = 11).value = numpy.amin(zAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 4).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        zAxis.insert(i, sheet.cell(row = writeRowNum, column = 4).value)
        i += 1

findXMin()
findYMin()
findZMin()

workbook.save(filename = dest_filename)
