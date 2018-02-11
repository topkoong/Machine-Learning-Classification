# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os
import numpy

os.chdir('browsing/User7_b_1499972368149')
dest_filename = 'dataToBeCalculated.xlsx'
#openpyxl.load_workbook(filename) returns a Workbook object.
workbook = openpyxl.load_workbook(filename = dest_filename)
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
sheet = workbook.get_sheet_by_name('dataToBeCalculated')

def findXMax():
    print("Finding Max: X Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    xAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nX Axis Max: %f" % numpy.amax(xAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 5, column = 12).value = numpy.amax(xAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 2).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        xAxis.insert(i, sheet.cell(row = writeRowNum, column = 2).value)
        i += 1

def findYMax():
    print("Finding Max: Y Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    yAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nY Axis Max: %f" % numpy.amax(yAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 5, column = 13).value = numpy.amax(yAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 3).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        yAxis.insert(i, sheet.cell(row = writeRowNum, column = 3).value)
        i += 1

def findZMax():
    print("Finding Max: Z Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 38235
    zAxis = []
    i = 0
    while (readRowNum <= maxRow):
        if(counter > 5):
            counter = 1
            print("\n\nZ Axis Max: %f" % numpy.amax(zAxis))
            print("=====================================\n")
            sheet.cell(row = writeRowNum - 5, column = 14).value = numpy.amax(zAxis)

        print("Reading from RowNo.", readRowNum," :", sheet.cell(row = readRowNum, column = 4).value)
        readRowNum += 1
        writeRowNum += 1
        counter += 1
        zAxis.insert(i, sheet.cell(row = writeRowNum, column = 4).value)
        i += 1

findXMax()
findYMax()
findZMax()

workbook.save(filename = dest_filename)
