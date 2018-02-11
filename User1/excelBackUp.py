# The OpenPyXL third-party module handles Excel spreadsheets (.xlsx files).
import openpyxl
import os

os.chdir('User1/browsing/USER1_b_1499548549244')
#openpyxl.load_workbook(filename) returns a Workbook object.
workbook = openpyxl.load_workbook('gyro_readings.xlsx')
wb = openpyxl.Workbook()
dest_filename = 'dataToBeCalculated.xlsx'
ws = wb.create_sheet(title="dataToBeCalculated")
#get_sheet__names() and get_sheet_by_name help get Worksheet objects.
sheet = workbook.get_sheet_by_name('gyro_readings')

# The cell() method also returns a Cell object from a sheet.


def getXAxisFromStarting():
    print("Reading Gyroscope: X Axis from Starting")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 100
    while (readRowNum <= maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1

        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 2).value)
        #x = sheet.cell(row = rowNum, column = 2).value
        ws.cell(row = writeRowNum, column = 2).value = sheet.cell(row = readRowNum, column = 2).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1

def getXAxisFromTheEnd():
    print("\nReading Gyroscope: X Axis from the End")
    print("=====================================\n")
    counter = 1
    writeRowNum = 165
    maxRow = 38335
    minRow = 38235
    readRowNum = maxRow
    while (readRowNum >= minRow):
        if(counter > 5):
            readRowNum += 2
            counter = 1
        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 2).value)
        ws.cell(row = writeRowNum, column = 2).value = sheet.cell(row = readRowNum, column = 2).value
        readRowNum -= 1
        writeRowNum += 1
        counter += 1



def getYAxisFromStarting():
    print("\nReading Gyroscope: Y Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 105
    while (readRowNum < maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1
        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 3).value)
        #x = sheet.cell(row = rowNum, column = 2).value
        ws.cell(row = writeRowNum, column = 3).value = sheet.cell(row = readRowNum, column = 3).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1

def getYAxisFromTheEnd():
    print("\nReading Gyroscope: Y Axis from the End")
    print("=====================================\n")
    counter = 1
    writeRowNum = 165
    maxRow = 38335
    minRow = 38235
    readRowNum = maxRow
    while (readRowNum >= minRow):
        if(counter > 5):
            readRowNum += 2
            counter = 1
        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 3).value)
        ws.cell(row = writeRowNum, column = 3).value = sheet.cell(row = readRowNum, column = 3).value
        readRowNum -= 1
        writeRowNum += 1
        counter += 1


def getZAxisFromStarting():
    print("\nReading Gyroscope: Z Axis")
    print("=====================================\n")
    counter = 1
    readRowNum = 1
    writeRowNum = 1
    maxRow = 105
    while (readRowNum < maxRow):
        if(counter > 5):
            readRowNum -= 2
            counter = 1

        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 4).value)
        ws.cell(row = writeRowNum, column = 4).value = sheet.cell(row = readRowNum, column = 4).value
        readRowNum += 1
        writeRowNum += 1
        counter += 1

def getZAxisFromTheEnd():
    print("\nReading Gyroscope: Z Axis from the End")
    print("=====================================\n")
    counter = 1
    writeRowNum = 165
    maxRow = 38335
    minRow = 38235
    readRowNum = maxRow
    while (readRowNum >= minRow):
        if(counter > 5):
            readRowNum += 2
            counter = 1
        print(writeRowNum, ": Reading from RowNo.", readRowNum," :", sheet.cell(row = writeRowNum, column = 4).value)
        ws.cell(row = writeRowNum, column = 4).value = sheet.cell(row = readRowNum, column = 4).value
        readRowNum -= 1
        writeRowNum += 1
        counter += 1


getXAxisFromStarting()
getXAxisFromTheEnd()
getYAxisFromStarting()
getYAxisFromTheEnd()
getZAxisFromStarting()
getZAxisFromTheEnd()
wb.save(filename = dest_filename)
